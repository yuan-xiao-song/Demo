"""
======================
-*- coding :utf-8 -*-
@File:excel.py
@Create:2020/05/09_14:43
@Email:18523607242@163.com
@Author: 袁晓松
@Software: PyCharm
======================
"""
import openpyxl  # python处理.xlsx 格式文件的三方库

from common.log import log

"""
excel文件处理
"""


class Excel:
    # 初始化excel类
    def __init__(self, excel_name, sheet_name):
        """
        传入excel文件名,表单名
        :param excel_name: excel文件名
        :param sheet_name: 表单名
        """
        # excel文件名，文件的绝对路径
        self.excel_name = excel_name
        # excel文件的表单名
        self.sheet_name = sheet_name

    def open_excel(self):
        """
        打开excel文件，选取excel文件中表单
        :return:None
        """
        # 加载excel文件为工作簿对象
        self.wb = openpyxl.load_workbook(self.excel_name)  # load_workbook 加载工作簿对象
        # 加载excel文件的表单为表单对象
        self.sh = self.wb[self.sheet_name]

    def get_data(self):
        """
        获取表单所有的数据
        :return: 测试用例列表
        """
        self.open_excel()
        # 读取表单中所有数据，并转换成为列表形式保存
        datas = list(self.sh.rows)  # rows 表单中的所有的行
        # new一个空列表接收表单中的测试用例数据
        cases = []
        # 读取datas中索引为0的所有数据，以列表形式保存
        title = [i.value for i in datas[0]]
        # 读取测试用例数据
        for j in datas[1:]:
            value = [k.value for k in j]
            # 组合测试用例，zip 聚合打包[(),(),()],dict 创建字典
            case = dict(zip(title, value))
            cases.append(case)  # 新增到列表cases中
        return cases

    def write_data(self, row, column, value=None):
        """
        写入数据到excel表格中的指定位置
        :param row:行
        :param column:lie
        :param value: 值
        :return: None
        """
        self.open_excel()
        # 写入数据到指定位置
        self.sh.cell(row=row, column=column, value=value)
        # 保存到文件 ---> ctrl + s
        self.wb.save(self.excel_name)


if __name__ == '__main__':
    # 实例化excel对象 第一个参数：excle文件的绝对路径,第二个参数，表单名
    res = Excel(r"D:\PycharmProjects\weixin\data\interfaceCase.xlsx", "login")
    # 读取表单中的数据，组合成为测试用例
    cases = res.get_data()
